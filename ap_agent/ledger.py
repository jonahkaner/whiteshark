"""Excel ledger for logging all AP transactions."""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

from ap_agent.invoice_parser import PaymentRequest

LEDGER_HEADERS = [
    "Date",
    "Vendor Name",
    "Vendor Email",
    "Invoice/Reference",
    "Amount",
    "Routing Number",
    "Account Number",
    "Bank Name",
    "Payment Terms",
    "Due Date",
    "Status",
    "Wire Confirmation #",
    "Approval Date",
    "Wire Sent Date",
    "Confirmation Received Date",
    "Vendor Notified Date",
    "Notes",
]


class Ledger:
    """Manage the AP Excel ledger."""

    def __init__(self, ledger_path: str):
        self.path = ledger_path
        self._ensure_file()

    def _ensure_file(self):
        """Create the ledger file with headers if it doesn't exist."""
        if not os.path.exists(self.path):
            wb = Workbook()
            ws = wb.active
            ws.title = "AP Ledger"
            ws.append(LEDGER_HEADERS)

            # Set column widths for readability
            widths = [12, 25, 30, 20, 15, 15, 18, 25, 15, 12, 15, 20, 12, 12, 12, 12, 30]
            for i, width in enumerate(widths, 1):
                ws.column_dimensions[
                    ws.cell(row=1, column=i).column_letter
                ].width = width

            wb.save(self.path)

    def log_payment_request(self, request: PaymentRequest) -> int:
        """Log a new payment request. Returns the row number."""
        wb = load_workbook(self.path)
        ws = wb.active

        row_data = [
            datetime.now().strftime("%Y-%m-%d"),
            request.vendor_name,
            request.vendor_email,
            request.invoice_reference or "",
            request.amount,
            request.routing_number,
            request.account_number,
            request.bank_name or "",
            request.payment_terms or "",
            request.due_date.isoformat() if request.due_date else "",
            "PENDING_APPROVAL",
            "",  # confirmation number
            "",  # approval date
            "",  # wire sent date
            "",  # confirmation received date
            "",  # vendor notified date
            "",  # notes
        ]
        ws.append(row_data)
        row_num = ws.max_row
        wb.save(self.path)
        return row_num

    # Column indices (1-based) — keep in sync with LEDGER_HEADERS
    COL_STATUS = 11
    COL_CONFIRMATION = 12
    COL_APPROVAL_DATE = 13
    COL_WIRE_SENT_DATE = 14
    COL_CONFIRM_DATE = 15
    COL_VENDOR_NOTIFIED = 16
    COL_NOTES = 17

    def update_status(self, row: int, status: str, notes: str = ""):
        """Update the status of a ledger entry."""
        wb = load_workbook(self.path)
        ws = wb.active
        ws.cell(row=row, column=self.COL_STATUS, value=status)
        if notes:
            ws.cell(row=row, column=self.COL_NOTES, value=notes)
        wb.save(self.path)

    def update_approval(self, row: int):
        """Mark a row as approved."""
        wb = load_workbook(self.path)
        ws = wb.active
        ws.cell(row=row, column=self.COL_STATUS, value="APPROVED")
        ws.cell(row=row, column=self.COL_APPROVAL_DATE,
                value=datetime.now().strftime("%Y-%m-%d %H:%M"))
        wb.save(self.path)

    def update_wire_sent(self, row: int):
        """Mark a row as wire sent."""
        wb = load_workbook(self.path)
        ws = wb.active
        ws.cell(row=row, column=self.COL_STATUS, value="WIRE_SENT")
        ws.cell(row=row, column=self.COL_WIRE_SENT_DATE,
                value=datetime.now().strftime("%Y-%m-%d %H:%M"))
        wb.save(self.path)

    def update_confirmation(self, row: int, confirmation_number: str):
        """Record the wire confirmation number."""
        wb = load_workbook(self.path)
        ws = wb.active
        ws.cell(row=row, column=self.COL_STATUS, value="CONFIRMED")
        ws.cell(row=row, column=self.COL_CONFIRMATION, value=confirmation_number)
        ws.cell(row=row, column=self.COL_CONFIRM_DATE,
                value=datetime.now().strftime("%Y-%m-%d %H:%M"))
        wb.save(self.path)

    def update_vendor_notified(self, row: int):
        """Mark that the vendor has been notified of the wire confirmation."""
        wb = load_workbook(self.path)
        ws = wb.active
        ws.cell(row=row, column=self.COL_STATUS, value="COMPLETED")
        ws.cell(row=row, column=self.COL_VENDOR_NOTIFIED,
                value=datetime.now().strftime("%Y-%m-%d %H:%M"))
        wb.save(self.path)

    def find_rows_by_status(self, status: str) -> list[dict]:
        """Find all rows with a given status. Returns list of dicts."""
        wb = load_workbook(self.path)
        ws = wb.active
        results = []
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=self.COL_STATUS).value == status:
                results.append({
                    "row": row_num,
                    "vendor_name": ws.cell(row=row_num, column=2).value,
                    "vendor_email": ws.cell(row=row_num, column=3).value,
                    "invoice_reference": ws.cell(row=row_num, column=4).value,
                    "amount": ws.cell(row=row_num, column=5).value,
                    "routing_number": ws.cell(row=row_num, column=6).value,
                    "account_number": ws.cell(row=row_num, column=7).value,
                    "bank_name": ws.cell(row=row_num, column=8).value,
                    "due_date": ws.cell(row=row_num, column=10).value,
                    "confirmation_number": ws.cell(
                        row=row_num, column=self.COL_CONFIRMATION
                    ).value,
                })
        return results
